# -*- coding: utf-8 -*-
"""
Created on Sat Apr  6 23:32:46 2019

@ author: xdong
@ Attention:
@ need openpyxl version > 2.6 
@ pip install --upgrade openpyxl
"""
from openpyxl import load_workbook, Workbook      #读取excel需要的模块
from openpyxl.styles import Border, Side, Font, Alignment,PatternFill,NamedStyle
import os

### 获取输出excel示例文件
def get_demo_sheet():    
    workbook=load_workbook("demo.xlsx")
    worksheet=workbook.worksheets[0]
    return worksheet
demosheet = get_demo_sheet()

def listfile(dirname,postfix = ''):
    filelist = []
    files = os.listdir(dirname)
    for item in files:
        #filelist.append([dirname,item])
        if os.path.isfile(dirname+item):
            if item.endswith(postfix):
                filelist.append([dirname,item])
        else:
            if os.path.isdir(dirname+item):
                filelist.extend(listfile(dirname+item+'\\',postfix))
    return filelist

### 获取输入excel文件内容
def find_in_sheet(sheet):
    for column in sheet.iter_cols():
        for cell2 in column:
            if cell2.value == "字段名称":
                return cell2.row, cell2.col_idx
    return -1, -1

bd = Side(style='thin', color="FF000000")
highlight = NamedStyle(name="highlight")
highlight.font = Font(name='Arial',size=10,bold=True,italic=False,\
           vertAlign=None,underline='none',\
           strike=False,color='FF000000') #字体
highlight.fill = PatternFill("solid", fgColor="00ffff") #背景填充
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd) #边框
highlight.alignment = Alignment(horizontal='center', vertical='center') #居中

highlight2 = NamedStyle(name="highlight2")
highlight2.font = Font(name='宋体',size=11,bold=True,italic=False,\
           vertAlign=None,underline='none',\
           strike=False,color='a020f0')
highlight2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight2.alignment = Alignment(horizontal='center', vertical='center')

highlight3 = NamedStyle(name="highlight2")
highlight3.font = Font(name='宋体',size=11,bold=True,italic=False,\
           vertAlign=None,underline='none',\
           strike=False,color='FF000000')
highlight3.border = Border(left=bd, top=bd, right=bd, bottom=bd)

def convertxlsx(infile,outfile):
    new_wb = Workbook()
    
    wb = load_workbook(infile)
    all_sheets = wb.get_sheet_names()
    for isheet in all_sheets:
        sheet = wb.get_sheet_by_name(isheet)
        srow, scol = find_in_sheet(sheet)
        if(srow != -1):
            new_sheet = new_wb.create_sheet(isheet)
            new_sheet.merge_cells('B1:C1')
            new_sheet.merge_cells('B2:C2')
            new_sheet.cell(row=1,column=1).value = demosheet.cell(row=1,column=1).value
            new_sheet.cell(row=2,column=1).value = demosheet.cell(row=2,column=1).value
            new_sheet.cell(row=3,column=1).value = demosheet.cell(row=3,column=1).value
            new_sheet.cell(row=1,column=2).value = demosheet.cell(row=1,column=2).value
            new_sheet.cell(row=2,column=2).value = demosheet.cell(row=2,column=2).value
            new_sheet.cell(row=3,column=2).value = demosheet.cell(row=3,column=2).value
            new_sheet.cell(row=3,column=3).value = demosheet.cell(row=3,column=3).value
            new_sheet.cell(row=1,column=1).style =highlight
            new_sheet.cell(row=2,column=1).style =highlight
            new_sheet.cell(row=3,column=1).style =highlight
            new_sheet.cell(row=1,column=2).style =highlight2
            new_sheet.cell(row=2,column=2).style =highlight2
            new_sheet.cell(row=1,column=3).style =highlight2
            new_sheet.cell(row=2,column=3).style =highlight2
            new_sheet.cell(row=3,column=2).style =highlight
            new_sheet.cell(row=3,column=3).style =highlight
            print(srow,scol,sheet.cell(row=srow-2, column=scol-1).value)
            try:
                engname = sheet.cell(row=srow-2, column=scol-1).value\
                .split('：')[1].strip('')
                new_sheet.cell(row=1, column=2).value = engname
            except Exception as e:
                print("Engname hasn't value!!!")
            try:
                chname  = sheet.cell(row=srow-1, column=scol-1).value\
                .split('：')[1].strip('')            
                new_sheet.cell(row=2, column=2).value = chname
            except Exception as e:
                print("Chname hasn't value!!!")
            crow = 4
            for irow in range(srow+1,sheet.max_row):
                iflag = 1
                try:
                    name = sheet.cell(row=irow, column=scol).value\
                            .strip().upper()
                    new_sheet.cell(row=crow,column=1).value = name
                    new_sheet.cell(row=crow,column=1).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    crow += 1
                    iflag = 0
                except Exception as e:
                    print("Name hasn't value!!!")
                try:
                    mean = sheet.cell(row=irow, column=scol+1).value\
                            .strip().upper()
                    new_sheet.cell(row=crow,column=2).value = mean               
                    new_sheet.cell(row=crow,column=2).border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if(iflag==1):      
                        crow += 1
                        print("Please check, Name hasn't value but mean does")
                except Exception as e:
                    print("Mean hasn't value!!!")
    new_wb.save(outfile)

## code debug
#infile = "D:\\mydata\\LY\\xie\\xie\\汇总层4\\111_汇总层数据字典.xlsx"
#outfile = "D:\\mydata\\LY\\xie\\xie\\new_111_汇总层数据字典.xlsx"
#convertxlsx(infile,outfile)
          
homedir = "D:\\mydata\\LY\\xie\\xie\\"
dirlist = ["汇总层4\\","中间层2\\","整合层1\\"]
for idir in dirlist:
    files = listfile(homedir+idir,postfix="")
    print(files)
    for ifile in files:
        infile = ifile[0] + ifile[1]
        outfile = ifile[0] + "new_" + ifile[1]
        convertxlsx(infile,outfile)
    

#from openpyxl.styles import Border, Side, Font, Alignment #设置字体和边框需要的模块
# 
#居中设置
#align = Alignment(horizontal='center', vertical='center')
#设置字体样式
#font = Font(name='宋体',size=11,bold=False,italic=False,\
#            vertAlign=None,underline='none',\
#            strike=False,color='FF000000')
##设置边框样式，上下左右边框
#border = Border(left=Side(style='medium',color='FF000000'),\
#                right=Side(style='medium',color='FF000000'),\
#                top=Side(style='medium',color='FF000000'),\
#                bottom=Side(style='medium',color='FF000000'),\
#                diagonal=Side(style='medium',color='FF000000'),\
#                diagonal_direction=0,\
#                outline=Side(style='medium',color='FF000000'),\
#                vertical=Side(style='medium',color='FF000000'),\
#                horizontal=Side(style='medium',color='FF000000'))
#os.chdir("c:\\bb")
#fpath_bb = os.getcwd() + "\\" + "鲁山报表.xlsx"
#print(fpath_bb)
#wbbb = load_workbook(fpath_bb)  #装载excel文件
#wsbbq = wbbb.get_sheet_by_name("南区")  #装载sheet表
#wsbbq["A35"] = "报表"  #excel 表格 赋值
#wsbbq['A35'].font=font  #设置单元格字体
#wsbbq['A38'].border=border #设置边框
